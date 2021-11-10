import time
import os
import sys
import requests
from openpyxl import load_workbook
import pandas as pd
import json
from threading import Thread, Lock


def readExcel(filename, sheetname):
    workbook = load_workbook(filename=filename)
    sheet = workbook[sheetname]
    _index = {}
    for i in range(1, sheet.max_column+1):
        _key = sheet.cell(row=1, column=i).value
        if _key not in _index.keys():
            _index[_key] = [i]
        else:
            _index[_key].append(i)
    data = []
    for i in range(2, sheet.max_row+1):
        dd = {}
        for key, value in _index.items():
            if len(value) == 1:
                dd[key] = sheet.cell(row=i, column=value[0]).value
            else:
                _value = []
                for ii in value:
                    _value.append(sheet.cell(row=i, column=ii).value)
                dd[key] = _value
        data.append(dd)
    return data


def writeExcel(df, filename, sheetname):
    if not os.path.exists(filename):
        tt = pd.DataFrame()
        tt.to_excel(filename)
    book = load_workbook(filename)
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    df.sort_values("index", inplace=True, ascending=True)
    df.to_excel(writer, sheetname)
    writer.save()
    return True


def find(target, _json):
    for key, value in _json.items():
        if key == target:
            return True
    return False


def T(target, data):
    vec = [data]
    for v in vec:
        for key, value in v.items():
            if isinstance(value, dict):
                if find(target=target, _json=value):
                    return key, value, "dict"
                else:
                    vec.append(value)
            if isinstance(value, list):
                if find(target=target, _json=value[0]):
                    return key, value, "list"
                else:
                    vec.append(value[0])
    return {}, {}, False


class Task(Thread):

    def __init__(self, cookie_str, url, method, tas, headers, name):
        Thread.__init__(self)
        cookies = {}
        for line in cookie_str.split(';'):
            key, value = line.split('=', 1)
            cookies[key] = value

        self.cookies = cookies
        self.url = url
        self.method = method
        self.tas = tas
        self.headers = headers
        self.name = name

    def run(self):
        global i
        global tbinfo
        lock.acquire()
        index = i
        i += 1
        lock.release()
        while index < end:
            print(self.name+'--'+str(index)+'--'+str(tb[index]))
            info = Task.getMetaData(tn=tb[index], self=self)  # 结果添加标签
            lock.acquire()
            if info != {}:
                k, v, t = T(target=tas, data=info)
                print(v)
                if t != False and t == 'list':
                    for vv in v:
                        vv["index"] = index
                        tbinfo.append(vv)
                if t != False and t == 'dict':
                    v["index"] = index
                    tbinfo.append(v)
            else:
                print('爬取第({})行的信息失败'.format(index))
            index = i
            i += 1
            lock.release()

    def getMetaData(self, tn):
        try:
            url = self.url
            method = self.method
            if method == "1":
                info = requests.post(url, headers=self.headers, cookies=self.cookies,
                                     data=tn).json()
            elif method == "2":
                info = requests.post(url, headers=self.headers, cookies=self.cookies,
                                     data=json.dumps(tn)).json()
            elif method == "3":
                url += '?'
                for k, v in tn.items():
                    url += str(k)+'='+str(v)+'&'
                url = url[:-1]
                info = requests.get(url, headers=self.headers,
                                    cookies=self.cookies).json()
            else:
                return {}
        except Exception as e:
            print(e)
            return {}
        return info


while True:
    url = input("请输入Request URL：")
    method = input(
        "请输入HTTP请求参数类型(1:Form Data 2:Request Payload 3:Query String Parameters)：")
    cookie = input("请输入浏览器复制的cookie值：")
    tas = input("请输入目标字段值：")
    if input("确认Request URL是{},Request Method是{}并且cookie和目标字段值{}正确,Y/N?".format(url, method, tas)) in {'Y', 'y', ''}:
        break

headers = {
    'User-agent': 'Mozilla/5.0 (Windows NT 6.1; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.113 Safari/537.36'}

while True:
    try:
        while True:
            tbpath = input("请输入参数所在文件路径，建议使用绝对路径：")
            st_name = input("请输入excel表格的sheetname，回车默认“Sheet1”：")
            if st_name == "":
                st_name = 'Sheet1'
            if input("确认路径是{}并且sheetname是{},Y/N?".format(tbpath, st_name)) in {'Y', 'y', ''}:
                break
        tb = readExcel(tbpath, st_name)
        target = input('输出第一行,检验是否输入正常,{},确认正常输入Y/y：'.format(str(tb[0])))
        if target in {'Y', 'y', ''}:
            break
    except Exception as e:
        print('读入参数文件有误')
        print(e)

i = 0
end = len(tb)
lock = Lock()
tbinfo = []

while True:
    tdnum = input("请输入线程的总数,建议5个左右,回车默认为1：")
    if tdnum == "":
        tdnum = 1
    if input("确认线程的总数是{},Y/N?".format(tdnum)) in {'Y', 'y', ''}:
        break

threads = []
for a in range(int(tdnum)):
    name = 'Thread-' + str(a)
    threads.append(
        Task(cookie_str=cookie, headers=headers, name=name, url=url, method=method, tas=tas))
for thread in threads:
    thread.start()

for t in threads:
    t.join()

while True:
    try:
        while True:
            filename = input(
                "信息爬取完毕,请输入文件名称,如:test.xlsx,路径与程序同一路径下,回车默认自动生成A_DATA_时间戳：")
            if getattr(sys, 'frozen', False):
                application_path = os.path.dirname(sys.executable)
            elif __file__:
                application_path = os.path.dirname(__file__)
            if filename == "":
                filename = "{}/A_DATA_{}.xlsx".format(
                    application_path, str(int(time.time())))
            else:
                filename = "{}/{}".format(application_path, filename)
            sheetname = input("请输入excel表格的sheetname，回车默认“Sheet1”：")
            if sheetname == "":
                sheetname = 'Sheet1'
            if input("确认路径是{}并且sheetname是{},Y/N?".format(filename, sheetname)) in {'Y', 'y', ''}:
                break

        if writeExcel(pd.DataFrame(tbinfo), filename=filename, sheetname=sheetname) == True:
            break
    except Exception as e:
        print('写入信息失败')
        print(e)
